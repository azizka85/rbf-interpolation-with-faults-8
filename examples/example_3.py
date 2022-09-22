from openpyxl import load_workbook

def load_data():
  wb = load_workbook('test.xlsx')

  data_sheet = wb.worksheets[0]
  fracture_sheet = wb.worksheets[1]

  points = []
  fault_data = []

  min_x = min_y = max_x = max_y = None

  for i in range(2, 76):
    x = data_sheet.cell(row=i, column=2).value
    y = data_sheet.cell(row=i, column=3).value
    z = data_sheet.cell(row=i, column=4).value

    min_x = min(min_x, x) if min_x != None else x
    max_x = max(max_x, x) if max_x != None else x

    min_y = min(min_y, y) if min_y != None else y
    max_y = max(max_y, y) if max_y != None else y

    points.append((x, y, z))

  for i in range(1, 8):
    x = fracture_sheet.cell(row=i, column=1).value
    y = fracture_sheet.cell(row=i, column=2).value

    fault_data.append((x, y))

  n = len(fault_data)

  faults = []

  for i in range(1, n):
    faults.append((
      fault_data[i-1],
      fault_data[i]
    ))

  return min_x, min_y, max_x, max_y, points, faults
